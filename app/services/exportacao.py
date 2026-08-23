import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import models
from app.services.prisma import calcular_fluxo_prisma


def exportar_projeto_excel(
    db: Session, projeto_id: int, caminho_saida: str
) -> str:
    """Gera um arquivo Excel completo (.xlsx) com múltiplas abas estilizadas

    reproduzindo os formulários de coleta e estatísticas do PRISMA 2020.
    """
    wb = openpyxl.Workbook()
    # Remove a aba padrão
    wb.remove(wb.active)

    # Estilos Visuais
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # -------------------------------------------------------------
    # 1. ABA PRISMA 2020
    # -------------------------------------------------------------
    ws_prisma = wb.create_sheet(title="📑 PRISMA 2020")
    ws_prisma.append(["Fluxo de Seleção de Estudos - PRISMA 2020"])
    ws_prisma.cell(row=1, column=1).font = title_font
    ws_prisma.append([])

    dados_prisma = calcular_fluxo_prisma(db, projeto_id)

    ws_prisma.append(["Etapa do PRISMA", "Quantidade de Artigos"])
    for cell in ws_prisma[3]:
        cell.fill = header_fill
        cell.font = header_font

    linhas_prisma = [
        ("Total de estudos identificados nas bases", dados_prisma["total_encontrados"]),
        ("Estudos duplicados removidos", dados_prisma["duplicados"]),
        ("Estudos triados (Título e Resumo)", dados_prisma["apos_duplicados"]),
        ("Estudos excluídos na triagem", dados_prisma["excluidos_triagem"]),
        ("Estudos recuperados para leitura completa", dados_prisma["texto_completo_avaliado"]),
        ("Estudos excluídos na leitura completa", dados_prisma["excluidos_texto_completo"]),
        ("ESTUDOS INCLUÍDOS NA REVISÃO FINAL", dados_prisma["estudos_incluidos"]),
    ]

    for label, qtd in linhas_prisma:
        ws_prisma.append([label, qtd])

    # -------------------------------------------------------------
    # 2. ABA ARTIGOS
    # -------------------------------------------------------------
    ws_artigos = wb.create_sheet(title="📊 Artigos")
    headers_artigos = [
        "ID",
        "Título",
        "Autores",
        "Ano",
        "Fonte",
        "DOI",
        "URL",
        "Status",
    ]
    ws_artigos.append(headers_artigos)

    for cell in ws_artigos[1]:
        cell.fill = header_fill
        cell.font = header_font

    artigos = (
        db.query(models.Artigo)
        .filter(models.Artigo.projeto_id == projeto_id)
        .all()
    )
    for a in artigos:
        ws_artigos.append(
            [a.id, a.titulo, a.autores, a.ano, a.fonte, a.doi, a.url, a.status]
        )

    # -------------------------------------------------------------
    # 3. ABA TRIAGEM
    # -------------------------------------------------------------
    ws_triagem = wb.create_sheet(title="📋 Triagem")
    headers_triagem = [
        "ID Artigo",
        "Título",
        "Decisão",
        "Motivo Exclusão",
        "Sugestão IA",
        "Confiança IA",
        "Revisor",
    ]
    ws_triagem.append(headers_triagem)

    for cell in ws_triagem[1]:
        cell.fill = header_fill
        cell.font = header_font

    triagens = (
        db.query(models.Triagem)
        .join(models.Artigo)
        .filter(models.Artigo.projeto_id == projeto_id)
        .all()
    )
    for t in triagens:
        ws_triagem.append(
            [
                t.artigo_id,
                t.artigo.titulo,
                t.decisao,
                t.motivo_exclusao,
                t.avaliacao_ia,
                f"{t.confianca_ia}%" if t.confianca_ia else "-",
                t.revisor,
            ]
        )

    # -------------------------------------------------------------
    # 4. ABA EXTRAÇÃO
    # -------------------------------------------------------------
    ws_extracao = wb.create_sheet(title="📚 Extração")
    headers_extracao = [
        "ID Artigo",
        "Título Artigo",
        "Código RQ",
        "Pergunta",
        "Resposta Encontrada",
        "Evidência / Trecho",
    ]
    ws_extracao.append(headers_extracao)

    for cell in ws_extracao[1]:
        cell.fill = header_fill
        cell.font = header_font

    extracoes = (
        db.query(models.Extracao)
        .join(models.Artigo)
        .filter(models.Artigo.projeto_id == projeto_id)
        .all()
    )
    for e in extracoes:
        ws_extracao.append(
            [
                e.artigo_id,
                e.artigo.titulo,
                e.questao.codigo,
                e.questao.pergunta,
                e.resposta,
                e.evidencia,
            ]
        )

    # Formatação Final: Ajustar largura das colunas e bordas em todas as abas
    for ws in wb.worksheets:
        ws.views.sheetView[0].showGridLines = True
        for row in ws.iter_rows():
            for cell in row:
                if cell.row > 1 and cell.fill != header_fill:
                    cell.font = data_font
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 3, 12), 50
            )

    wb.save(caminho_saida)
    return caminho_saida